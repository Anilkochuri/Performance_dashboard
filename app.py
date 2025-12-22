#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Performance Dashboard (Flask)
- Static logo served via Flask static path
- Health endpoint for uptime monitors
- Upload Excel (.xlsx/.xls) and parse into in-memory cache
- Optional support for GET ?csvFile=... (reads from uploads/ or static/data/)
- Export to Excel (openpyxl) and PDF (reportlab + matplotlib)
"""

import os
import io
import uuid
from types import SimpleNamespace
from typing import Dict, Any, List

from flask import (
    Flask, request, redirect, url_for, session, flash,
    render_template_string, send_file
)

import pandas as pd
import matplotlib
matplotlib.use('Agg')  # non-GUI backend for servers
import matplotlib.pyplot as plt

# ----------------------------------------------------------------------------
# App & Config
# ----------------------------------------------------------------------------
app = Flask(__name__, static_folder="static", static_url_path="/static")
app.secret_key = os.environ.get("FLASK_SECRET_KEY", "change-this-secret-key")
app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024  # 50 MB

ALLOWED_EXTENSIONS = {"xlsx", "xls"}
TOP_N = int(os.environ.get('TOP_N', '10'))

BASE_DIR = os.getcwd()
UPLOAD_DIR = os.path.join(BASE_DIR, "uploads")
DATA_DIR = os.path.join(BASE_DIR, "static", "data")
os.makedirs(UPLOAD_DIR, exist_ok=True)

# In-memory per-session cache
CACHE: Dict[str, Dict[str, Any]] = {}

# ----------------------------------------------------------------------------
# Inline Templates (you can move to templates/*.html later)
# ----------------------------------------------------------------------------
LOGIN_TEMPLATE = r"""
<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <title>Login - Performance Dashboard</title>
  <link rel="stylesheet" href="{{ url_for('static', filename='css/style.css') }}">
</head>
<body style="font-family:Arial, sans-serif;">
  <div style="max-width:480px;margin:40px auto;padding:20px;border:1px solid #ddd;border-radius:8px;">
    <div style="display:flex;align-items:center;gap:12px;">
      <img src="{{ url_for('static', filename='images/logo.png') }}" alt="Logo" style="height:32px;">
      <h2 style="margin:0;">Performance Dashboard</h2>
    </div>
    <hr/>
    {% with messages = get_flashed_messages() %}{% if messages %}
      <div style="color:#b00;margin-bottom:10px;">{{ messages[0] }}</div>
    {% endif %}{% endwith %}
    <form method="POST" action="{{ url_for('login') }}">
      <label>Username</label>
      <input type="text" name="username" style="width:100%;margin:6px 0;padding:8px;">
      <label>Password</label>
      <input type="password" name="password" style="width:100%;margin:6px 0;padding:8px;">
      <button type="submit" style="margin-top:10px;padding:10px 16px;">Login</button>
    </form>
  </div>
</body>
</html>
"""

DASHBOARD_TEMPLATE = r"""
<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <title>Performance Dashboard</title>
  <link rel="stylesheet" href="{{ url_for('static', filename='css/style.css') }}">
</head>
<body style="font-family:Arial, sans-serif;background:#e6f0fb;">
  <header style="background:#4ea3ff;color:#fff;padding:10px 16px;display:flex;align-items:center;justify-content:space-between;">
    <div style="display:flex;align-items:center;gap:12px;">
      <img src="{{ url_for('static', filename='images/logo.png') }}" alt="Logo" style="height:32px;">
      <strong>Dashboard</strong>
    </div>
    <nav style="display:flex;gap:12px;">
      <a href="{{ url_for('export_xlsx') }}" style="color:#fff;">Export XLSX</a>
      <a href="{{ url_for('export_pdf') }}" style="color:#fff;">Export PDF</a>
      <a href="{{ url_for('logout') }}" style="color:#fff;">Logout</a>
    </nav>
  </header>

  <main style="max-width:1000px;margin:16px auto;padding:0 16px;">
    {% with messages = get_flashed_messages() %}{% if messages %}
      <div style="color:#b00;margin-bottom:10px;">{{ messages[0] }}</div>
    {% endif %}{% endwith %}

    <section style="background:#fff;border:1px solid #ddd;border-radius:8px;padding:12px;">
      <h3>Upload XLSX</h3>
      <form action="{{ url_for('upload_xlsx') }}" method="POST" enctype="multipart/form-data">
        <input type="file" name="file" accept=".xlsx,.xls">
        <button type="submit">Upload</button>
      </form>
      <p style="margin:8px 0 0;">Or open a baked demo file:
        <a href="{{ url_for('dashboard') }}?csvFile=LoadTest_Result_UPDATED_USING_NW.xlsx">LoadTest_Result_UPDATED_USING_NW.xlsx</a>
      </p>
    </section>

    {% if not data_loaded %}
      <p style="margin-top:12px;">Upload an Excel (.xlsx/.xls) to populate Observations, Transaction Summary, DB_HITS, and charts.</p>
    {% endif %}

    <section style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:12px;margin-top:12px;">
      <div style="background:#fff;border:1px solid #ddd;border-radius:8px;padding:12px;">
        <h4>&lt; 5s</h4><div>{{ observations.lt_5s }}</div>
        <h4>5–10s</h4><div>{{ observations.between_5_10s }}</div>
        <h4>&gt; 10s</h4><div>{{ observations.gt_10s }}</div>
        <h4>Total Pass / Fail</h4><div>{{ observations.total_pass }} / {{ observations.total_fail }}</div>
      </div>

      <div style="grid-column: span 2;background:#fff;border:1px solid #ddd;border-radius:8px;padding:12px;">
        <h4>Transaction Summary</h4>
        <div style="overflow:auto;max-height:360px;">
          <table border="1" cellspacing="0" cellpadding="6" style="border-collapse:collapse;width:100%;">
            <tr>
              {% for col in ts_columns %}<th>{{ col }}</th>{% endfor %}
            </tr>
            {% for row in ts_rows %}
              <tr>
                {% for col in ts_columns %}<td>{{ row.get(col, '') }}</td>{% endfor %}
              </tr>
            {% endfor %}
          </table>
        </div>
      </div>
    </section>

    <section style="margin-top:12px;background:#fff;border:1px solid #ddd;border-radius:8px;padding:12px;">
      <h4>DB_HITS</h4>
      {% if db_columns %}
        <div style="overflow:auto;max-height:300px;">
          <table border="1" cellspacing="0" cellpadding="6" style="border-collapse:collapse;width:100%;">
            <tr>
              {% for col in db_columns %}<th>{{ col }}</th>{% endfor %}
            </tr>
            {% for row in db_rows %}
              <tr>
                {% for col in db_columns %}<td>{{ row.get(col, '') }}</td>{% endfor %}
              </tr>
            {% endfor %}
          </table>
        </div>
      {% else %}
        <em>DB_HITS sheet not found or no data.</em>
      {% endif %}
    </section>
  </main>

  <footer style="text-align:center;color:#666;margin:24px 0;">© 2024 Dashboard Application</footer>
</body>
</html>
"""

# ----------------------------------------------------------------------------
# Utilities
# ----------------------------------------------------------------------------
def allowed_file(filename: str) -> bool:
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


def read_excel_sheets(file_bytes: bytes, ext: str) -> Dict[str, pd.DataFrame]:
    bio = io.BytesIO(file_bytes)
    engine = "openpyxl" if ext == "xlsx" else "xlrd"
    xls = pd.ExcelFile(bio, engine=engine)
    sheets_available = set(xls.sheet_names)

    def safe_read(name: str) -> pd.DataFrame:
        return xls.parse(name) if name in sheets_available else pd.DataFrame()

    return {
        "Transaction Summary": safe_read("Transaction Summary"),
        "DB_HITS": safe_read("DB_HITS"),
        "Response Codes": safe_read("Response Codes") if "Response Codes" in sheets_available else pd.DataFrame(),
    }

DEF_COLS = {
    'txn': ['transaction name', 'transaction', 'txn', 'sampler', 'label'],
    'avg': ['average', 'avg', 'mean'],
    'p90': ['90 percent', '90%', 'p90', '90th percentile'],
    'pass': ['pass', 'passed'],
    'fail': ['fail', 'failed'],
}


def _pick_column(df: pd.DataFrame, keys) -> str:
    cols = [str(c).strip() for c in df.columns]
    lowered = [c.lower() for c in cols]
    # exact match first
    for key in keys:
        kl = key.lower()
        for c, lc in zip(cols, lowered):
            if lc == kl:
                return c
    # then contains
    for key in keys:
        kl = key.lower()
        for c, lc in zip(cols, lowered):
            if kl in lc:
                return c
    return ''


def compute_observations_from_ts(ts_df: pd.DataFrame) -> Dict[str, Any]:
    obs = {"lt_5s": 0, "between_5_10s": 0, "gt_10s": 0, "total_pass": 0, "total_fail": 0}
    if ts_df.empty:
        return obs
    df = ts_df.copy(); df.columns = [str(c).strip() for c in df.columns]
    avg_col = _pick_column(df, DEF_COLS['avg'])
    pass_col = _pick_column(df, DEF_COLS['pass'])
    fail_col = _pick_column(df, DEF_COLS['fail'])
    if avg_col:
        avg_vals = pd.to_numeric(df[avg_col], errors='coerce')
        obs['lt_5s'] = int((avg_vals < 5).sum())
        obs['between_5_10s'] = int(((avg_vals >= 5) & (avg_vals <= 10)).sum())
        obs['gt_10s'] = int((avg_vals > 10).sum())
    if pass_col:
        obs['total_pass'] = int(pd.to_numeric(df[pass_col], errors='coerce').fillna(0).sum())
    if fail_col:
        obs['total_fail'] = int(pd.to_numeric(df[fail_col], errors='coerce').fillna(0).sum())
    return obs


def build_chart_payloads(ts_df: pd.DataFrame, resp_df: pd.DataFrame, db_hits_df: pd.DataFrame) -> Dict[str, Any]:
    payload: Dict[str, Any] = {
        "avg_bar": {},
        "p90_bar": {},
        "throughput": {},
        "total_pass_fail_pie": {},
        "resp_code_pie": {},
        "latency_bands_pie": {},
        "txn_count_avg_combo": {},
        "topn_avg_bar": {},
    }
    if ts_df.empty:
        return payload

    df = ts_df.copy(); df.columns = [str(c).strip() for c in df.columns]
    txn_col = _pick_column(df, DEF_COLS['txn']) or df.columns[0]
    avg_col = _pick_column(df, DEF_COLS['avg']) or (df.columns[1] if len(df.columns) > 1 else '')
    p90_col = _pick_column(df, DEF_COLS['p90'])
    pass_col = _pick_column(df, DEF_COLS['pass'])
    fail_col = _pick_column(df, DEF_COLS['fail'])

    # Average bar
    if txn_col and avg_col:
        avg_vals = pd.to_numeric(df[avg_col], errors='coerce').fillna(0)
        payload['avg_bar'] = {
            'x': df[txn_col].astype(str).tolist(),
            'y': avg_vals.tolist(),
            'yaxis_range': [0, 10],
            'title': 'Average Response Time (s) by Transaction'
        }
        # Top-N slowest by Average
        top_df = df[[txn_col, avg_col]].copy()
        top_df[avg_col] = pd.to_numeric(top_df[avg_col], errors='coerce').fillna(0)
        top_df = top_df.sort_values(by=avg_col, ascending=False).head(TOP_N)
        payload['topn_avg_bar'] = {
            'x': top_df[txn_col].astype(str).tolist(),
            'y': top_df[avg_col].tolist(),
            'yaxis_range': [0, 10],
            'title': f'Top-{TOP_N} Slowest Transactions (by Average)'
        }

    # P90 bar
    if txn_col and p90_col:
        payload['p90_bar'] = {
            'x': df[txn_col].astype(str).tolist(),
            'y': pd.to_numeric(df[p90_col], errors='coerce').fillna(0).tolist(),
            'yaxis_range': [0, 10],
            'title': '90th Percentile Response Time (s) by Transaction'
        }

    # Throughput + totals pie
    if txn_col and pass_col and fail_col:
        pass_y = pd.to_numeric(df[pass_col], errors='coerce').fillna(0)
        fail_y = pd.to_numeric(df[fail_col], errors='coerce').fillna(0)
        x = df[txn_col].astype(str).tolist()
        payload['throughput'] = {
            'x': x,
            'pass': pass_y.tolist(),
            'fail': fail_y.tolist(),
            'title': 'Throughput: Pass & Fail counts per Transaction'
        }
        payload['total_pass_fail_pie'] = {
            'labels': ['Pass', 'Fail'],
            'values': [int(pass_y.sum()), int(fail_y.sum())]
        }
        if avg_col:
            payload['txn_count_avg_combo'] = {
                'x': x,
                'count': (pass_y + fail_y).tolist(),
                'avg': pd.to_numeric(df[avg_col], errors='coerce').fillna(0).tolist(),
                'title': 'Transactions vs Count and Average Response Time'
            }

    # Response Codes pie (optional)
    resp_pie = {"labels": [], "values": []}
    if not resp_df.empty:
        d = resp_df.copy(); d.columns = [str(c).strip() for c in d.columns]
        code_col = _pick_column(d, ['response code', 'code', 'response'])
        count_col = _pick_column(d, ['count', 'total'])
        if code_col and count_col:
            resp_pie['labels'] = d[code_col].astype(str).tolist()
            resp_pie['values'] = pd.to_numeric(d[count_col], errors='coerce').fillna(0).tolist()
            payload['resp_code_pie'] = resp_pie

    # Latency bands pie
    if avg_col:
        avg_vals = pd.to_numeric(df[avg_col], errors='coerce').fillna(0)
        bins = [0, 2, 3, 5, 10, float('inf')]
        labels = ['<2s', '2–3s', '3–5s', '5–10s', '>10s']
        cats = pd.cut(avg_vals, bins=bins, labels=labels, right=False)
        counts = cats.value_counts().reindex(labels).fillna(0).astype(int)
        payload['latency_bands_pie'] = {'labels': labels, 'values': counts.tolist(), 'title': 'Response Time Distribution (Average)'}

    return payload

# ----------------------------------------------------------------------------
# Routes
# ----------------------------------------------------------------------------
@app.route('/')
def home():
    if session.get('user'):
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
        if username == os.environ.get('DASH_USERNAME', 'admin') and password == os.environ.get('DASH_PASSWORD', 'password'):
            session['user'] = username
            session['sid'] = session.get('sid') or str(uuid.uuid4())
            return redirect(url_for('dashboard'))
        flash('Invalid credentials. Try admin / password (default).')
    return render_template_string(LOGIN_TEMPLATE)


@app.route('/logout')
def logout():
    sid = session.get('sid')
    if sid and sid in CACHE:
        CACHE.pop(sid, None)
    session.clear()
    return redirect(url_for('login'))


@app.route('/health')
def health():
    return {"status": "ok"}, 200


@app.route('/dashboard', methods=['GET'])
def dashboard():
    if not session.get('user'):
        return redirect(url_for('login'))

    sid = session.get('sid')
    cache = CACHE.get(sid, {}) if sid else {}

    # Optional: allow direct GET with ?csvFile=...
    csv_file = request.args.get("csvFile")
    if csv_file and not cache.get('ts_df'):
        candidates = [
            os.path.join(UPLOAD_DIR, csv_file),
            os.path.join(DATA_DIR, csv_file),
        ]
        for path in candidates:
            if os.path.exists(path):
                ext = os.path.splitext(path)[1].lower().replace('.', '')
                with open(path, 'rb') as fh:
                    dfs = read_excel_sheets(fh.read(), ext)
                ts_df = dfs.get('Transaction Summary', pd.DataFrame())
                db_df = dfs.get('DB_HITS', pd.DataFrame())
                resp_df = dfs.get('Response Codes', pd.DataFrame())
                observations = compute_observations_from_ts(ts_df)
                chart_data = build_chart_payloads(ts_df, resp_df, db_df)
                CACHE[sid] = {
                    'observations': observations,
                    'ts_rows': ts_df.fillna('').to_dict(orient='records') if not ts_df.empty else [],
                    'ts_columns': [str(c).strip() for c in ts_df.columns],
                    'db_rows': db_df.fillna('').to_dict(orient='records') if not db_df.empty else [],
                    'db_columns': [str(c).strip() for c in db_df.columns],
                    'chart_data': chart_data,
                    'ts_df': ts_df, 'db_df': db_df, 'resp_df': resp_df,
                }
                flash(f'Loaded {csv_file}')
                break

    observations = cache.get('observations', {"lt_5s":0,"between_5_10s":0,"gt_10s":0,"total_pass":0,"total_fail":0})
    ts_rows = cache.get('ts_rows', [])
    ts_columns = cache.get('ts_columns', [])
    db_rows = cache.get('db_rows', [])
    db_columns = cache.get('db_columns', [])
    chart_data = cache.get('chart_data', {})
    data_loaded = bool(ts_rows)

    return render_template_string(
        DASHBOARD_TEMPLATE,
        observations=SimpleNamespace(**observations),
        ts_rows=ts_rows, ts_columns=ts_columns,
        db_rows=db_rows, db_columns=db_columns,
        chart_data=chart_data, data_loaded=data_loaded,
    )


@app.route('/upload_xlsx', methods=['POST'])
def upload_xlsx():
    if not session.get('user'):
        return redirect(url_for('login'))

    if 'file' not in request.files:
        flash('No file part'); return redirect(url_for('dashboard'))
    file = request.files['file']
    if file.filename == '':
        flash('No selected file'); return redirect(url_for('dashboard'))
    if not allowed_file(file.filename):
        flash('Invalid file type. Please upload .xlsx or .xls'); return redirect(url_for('dashboard'))

    ext = file.filename.rsplit('.', 1)[1].lower()
    file_bytes = file.read()

    # Save to runtime uploads so ?csvFile=... can work after upload
    save_path = os.path.join(UPLOAD_DIR, file.filename)
    try:
        with open(save_path, 'wb') as fh:
            fh.write(file_bytes)
    except Exception:
        pass  # optional; continue even if save fails

    try:
        dfs = read_excel_sheets(file_bytes, ext)
        ts_df = dfs.get('Transaction Summary', pd.DataFrame())
        db_df = dfs.get('DB_HITS', pd.DataFrame())
        resp_df = dfs.get('Response Codes', pd.DataFrame())

        observations = compute_observations_from_ts(ts_df)
        ts_cols = [str(c).strip() for c in ts_df.columns]
        ts_rows = ts_df.fillna('').to_dict(orient='records') if not ts_df.empty else []
        db_cols = [str(c).strip() for c in db_df.columns]
        db_rows = db_df.fillna('').to_dict(orient='records') if not db_df.empty else []
        chart_data = build_chart_payloads(ts_df, resp_df, db_df)

        sid = session.get('sid') or str(uuid.uuid4()); session['sid'] = sid
        CACHE[sid] = {
            'observations': observations,
            'ts_rows': ts_rows, 'ts_columns': ts_cols,
            'db_rows': db_rows, 'db_columns': db_cols,
            'chart_data': chart_data,
            'ts_df': ts_df, 'db_df': db_df, 'resp_df': resp_df,
        }
        flash('Excel processed successfully.')
    except Exception as e:
        print('Upload XLSX error:', e)
        flash(f'Failed to process Excel: {e}')
    return redirect(url_for('dashboard'))


# ----------------------------------------------------------------------------
# Export: XLSX and PDF
# ----------------------------------------------------------------------------
@app.route('/export_xlsx', methods=['GET'])
def export_xlsx():
    if not session.get('user'):
        return redirect(url_for('login'))
    sid = session.get('sid'); cache = CACHE.get(sid)
    if not cache:
        flash('No data to export. Upload Excel first.'); return redirect(url_for('dashboard'))

    from openpyxl import Workbook
    from openpyxl.chart import BarChart, PieChart, LineChart, Reference

    wb = Workbook()
    # Observations
    ws_obs = wb.active; ws_obs.title = 'Observations'
    obs = cache.get('observations', {})
    ws_obs.append(['Metric', 'Count'])
    ws_obs.append(['<5s', obs.get('lt_5s', 0)])
    ws_obs.append(['5–10s', obs.get('between_5_10s', 0)])
    ws_obs.append(['>10s', obs.get('gt_10s', 0)])
    ws_obs.append(['Total Pass', obs.get('total_pass', 0)])
    ws_obs.append(['Total Fail', obs.get('total_fail', 0)])

    # Transaction Summary
    ws_ts = wb.create_sheet('Transaction Summary')
    ts_df: pd.DataFrame = cache.get('ts_df', pd.DataFrame())
    if not ts_df.empty:
        ws_ts.append(list(ts_df.columns))
        for row in ts_df.fillna('').itertuples(index=False):
            ws_ts.append(list(row))

    # DB_HITS
    ws_db = wb.create_sheet('DB_HITS')
    db_df: pd.DataFrame = cache.get('db_df', pd.DataFrame())
    if not db_df.empty:
        ws_db.append(list(db_df.columns))
        for row in db_df.fillna('').itertuples(index=False):
            ws_db.append(list(row))

    # Response Codes
    ws_rc = wb.create_sheet('Response Codes')
    resp_df: pd.DataFrame = cache.get('resp_df', pd.DataFrame())
    if not resp_df.empty:
        ws_rc.append(list(resp_df.columns))
        for row in resp_df.fillna('').itertuples(index=False):
            ws_rc.append(list(row))

    # Various Graphs (data + charts)
    ws_vg = wb.create_sheet('Various Graphs')
    chart_data = cache.get('chart_data', {})
    ws_vg.append(['Transaction Name', 'Average', 'P90', 'Pass', 'Fail', 'Total'])
    ts_df_local = cache.get('ts_df', pd.DataFrame())
    if not ts_df_local.empty:
        txn_col = _pick_column(ts_df_local, DEF_COLS['txn']) or ts_df_local.columns[0]
        avg_col = _pick_column(ts_df_local, DEF_COLS['avg'])
        p90_col = _pick_column(ts_df_local, DEF_COLS['p90'])
        pass_col = _pick_column(ts_df_local, DEF_COLS['pass'])
        fail_col = _pick_column(ts_df_local, DEF_COLS['fail'])
        for _, r in ts_df_local.iterrows():
            tn = str(r.get(txn_col, ''))
            av = r.get(avg_col, 0)
            p9 = r.get(p90_col, 0)
            ps = r.get(pass_col, 0)
            fl = r.get(fail_col, 0)
            try:
                tot = float(pd.to_numeric(pd.Series([ps]), errors='coerce').fillna(0)[0]) + \
                      float(pd.to_numeric(pd.Series([fl]), errors='coerce').fillna(0)[0])
            except Exception:
                tot = 0
            ws_vg.append([tn, av, p9, ps, fl, tot])
        last_row = ws_vg.max_row
        cats = Reference(ws_vg, min_col=1, min_row=2, max_row=last_row)
        # Average bar
        bar = BarChart(); bar.title = 'Average Response Time (s) by Transaction'; bar.y_axis.title='Seconds'; bar.x_axis.title='Transaction Name'
        data = Reference(ws_vg, min_col=2, min_row=1, max_row=last_row)
        bar.add_data(data, titles_from_data=True, from_rows=False)
        bar.set_categories(cats)
        ws_vg.add_chart(bar, 'H2')
        # P90 bar
        pbar = BarChart(); pbar.title='90th Percentile Response Time (s)'; pbar.y_axis.title='Seconds'; pbar.x_axis.title='Transaction Name'
        pdata = Reference(ws_vg, min_col=3, min_row=1, max_row=last_row)
        pbar.add_data(pdata, titles_from_data=True, from_rows=False)
        pbar.set_categories(cats)
        ws_vg.add_chart(pbar, 'H20')
        # Throughput stacked bar
        tbar = BarChart(); tbar.type='col'; tbar.grouping='stacked'; tbar.title='Throughput: Pass & Fail per Transaction'; tbar.y_axis.title='Count'; tbar.x_axis.title='Transaction Name'
        pass_ref = Reference(ws_vg, min_col=4, min_row=1, max_row=last_row)
        fail_ref = Reference(ws_vg, min_col=5, min_row=1, max_row=last_row)
        tbar.add_data(pass_ref, titles_from_data=True, from_rows=False)
        tbar.add_data(fail_ref, titles_from_data=True, from_rows=False)
        tbar.set_categories(cats)
        ws_vg.add_chart(tbar, 'H38')
        # Combo Total vs Average
        combo_bar = BarChart(); combo_bar.title='Total Transactions'; combo_bar.y_axis.title='Count'; combo_bar.x_axis.title='Transaction Name'
        total_ref = Reference(ws_vg, min_col=6, min_row=1, max_row=last_row)
        combo_bar.add_data(total_ref, titles_from_data=True, from_rows=False)
        combo_bar.set_categories(cats)
        ws_vg.add_chart(combo_bar, 'H56')
        avg_line = LineChart(); avg_line.title='Average (s)'; avg_line.y_axis.title='Seconds'
        avg_ref = Reference(ws_vg, min_col=2, min_row=1, max_row=last_row)
        avg_line.add_data(avg_ref, titles_from_data=True, from_rows=False)
        avg_line.set_categories(cats)
        ws_vg.add_chart(avg_line, 'O56')
        # Latency bands pie
        lb_labels = chart_data.get('latency_bands_pie', {}).get('labels', [])
        lb_values = chart_data.get('latency_bands_pie', {}).get('values', [])
        if lb_labels and lb_values:
            ws_vg.append([]); ws_vg.append(['Latency Band', 'Count'])
            start = ws_vg.max_row
            for lbl, val in zip(lb_labels, lb_values):
                ws_vg.append([lbl, val])
            end = ws_vg.max_row
            pie = PieChart(); pie.title = 'Response Time Distribution (Average)'
            pie_data = Reference(ws_vg, min_col=2, min_row=start, max_row=end)
            pie_labels = Reference(ws_vg, min_col=1, min_row=start, max_row=end)
            pie.add_data(pie_data, titles_from_data=False)
            pie.set_categories(pie_labels)
            ws_vg.add_chart(pie, 'H74')

    bio = io.BytesIO(); wb.save(bio); bio.seek(0)
    return send_file(bio, as_attachment=True,
                     download_name='Performance_Dashboard_Export.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/export_pdf', methods=['GET'])
def export_pdf():
    if not session.get('user'):
        return redirect(url_for('login'))
    sid = session.get('sid'); cache = CACHE.get(sid)
    if not cache:
        flash('No data to export. Upload Excel first.'); return redirect(url_for('dashboard'))

    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image, PageBreak, Table, TableStyle
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_LEFT
    from reportlab.lib import colors

    styles = getSampleStyleSheet()
    cell_style = ParagraphStyle('Cell', parent=styles['Normal'], fontSize=8, leading=9, alignment=TA_LEFT)

    bio = io.BytesIO()
    doc = SimpleDocTemplate(bio, pagesize=landscape(A4))
    elements: List[Any] = []

    def add_title(t):
        elements.append(Paragraph(f"<b>{t}</b>", styles['Heading2']))
        elements.append(Spacer(1, 8))

    def fig_to_image(fig, w=700, h=380):
        buf = io.BytesIO()
        fig.savefig(buf, format='png', bbox_inches='tight', dpi=200)
        plt.close(fig)
        buf.seek(0)
        return Image(buf, width=w, height=h)

    # Title + Observations
    obs = cache.get('observations', {})
    elements.append(Paragraph('<b>Performance Dashboard</b>', styles['Title']))
    elements.append(Spacer(1, 12))
    add_title('Observations')
    elements.append(Paragraph(
        f"<b>&lt; 5s:</b> {obs.get('lt_5s',0)} &nbsp; "
        f"<b>5–10s:</b> {obs.get('between_5_10s',0)} &nbsp; "
        f"<b>&gt;10s:</b> {obs.get('gt_10s',0)} &nbsp; "
        f"<b>Total Pass/Fail:</b> {obs.get('total_pass',0)} / {obs.get('total_fail',0)}",
        styles['BodyText']
    ))
    fig, ax = plt.subplots(figsize=(5.5, 4))
    ax.pie([obs.get('total_pass',0), obs.get('total_fail',0)], labels=['Pass','Fail'], autopct='%1.0f%%',
           colors=['#4daf4a','#e41a1c'])
    ax.set_title('Total Transactions: Pass vs Fail')
    elements.append(fig_to_image(fig, w=360, h=240))
    elements.append(PageBreak())

    # Transaction Summary table
    ts_df: pd.DataFrame = cache.get('ts_df', pd.DataFrame())
    if not ts_df.empty:
        add_title('Transaction Summary (all rows)')
        header = [str(c) for c in ts_df.columns]
        def make_cell(v, max_len=200):
            s = '' if pd.isna(v) else str(v)
            if len(s) > max_len:
                s = s[:max_len] + '…'
            return Paragraph(s, cell_style)
        data_rows = ts_df.fillna('').applymap(lambda v: make_cell(v, 200)).values.tolist()
        table_data = [header] + data_rows
        sample = ts_df.head(200).copy().astype(str)
        lengths = [max(len(h), int(sample[col].map(lambda x: len(str(x))).mean())) for h, col in zip(header, ts_df.columns)]
        total_len = max(sum(lengths), 1)
        usable_width = 780
        col_widths = [max(60, int(usable_width * (l / total_len))) for l in lengths]
        table = Table(table_data, colWidths=col_widths, repeatRows=1)
        table.setStyle(TableStyle([
            ('GRID', (0,0), (-1,-1), 0.25, colors.grey),
            ('BACKGROUND', (0,0), (-1,0), colors.whitesmoke),
            ('FONTSIZE', (0,0), (-1,-1), 8),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ]))
        elements.append(table)
        elements.append(PageBreak())

    # Charts
    chart_data = cache.get('chart_data', {})
    ts_df: pd.DataFrame = cache.get('ts_df', pd.DataFrame())
    if not ts_df.empty:
        df = ts_df.copy(); df.columns = [str(c).strip() for c in df.columns]
        txn_col = _pick_column(df, DEF_COLS['txn']) or df.columns[0]
        avg_col = _pick_column(df, DEF_COLS['avg'])
        p90_col = _pick_column(df, DEF_COLS['p90'])
        pass_col = _pick_column(df, DEF_COLS['pass'])
        fail_col = _pick_column(df, DEF_COLS['fail'])
        x = df[txn_col].astype(str).tolist()
        avg = pd.to_numeric(df[avg_col], errors='coerce').fillna(0).tolist() if avg_col else []
        p90 = pd.to_numeric(df[p90_col], errors='coerce').fillna(0).tolist() if p90_col else []
        ps = pd.to_numeric(df[pass_col], errors='coerce').fillna(0).tolist() if pass_col else []
        fl = pd.to_numeric(df[fail_col], errors='coerce').fillna(0).tolist() if fail_col else []

        def downsample_ticks(names: List[str], max_ticks: int = 25):
            n = len(names)
            if n <= max_ticks: return list(range(n)), names
            step = max(1, n // max_ticks)
            idxs = list(range(0, n, step))
            return idxs, [names[i] for i in idxs]

        if avg:
            fig, ax = plt.subplots(figsize=(10, 6))
            ax.bar(range(len(x)), avg, color='#2b8cbe')
            ticks, labels = downsample_ticks(x)
            ax.set_xticks(ticks); ax.set_xticklabels(labels, rotation=45, ha='right')
            ax.set_title('Average Response Time (s) by Transaction')
            ax.set_ylabel('Seconds'); ax.set_ylim(0, 10)
            elements.append(fig_to_image(fig))

        if p90:
            fig, ax = plt.subplots(figsize=(10, 6))
            ax.bar(range(len(x)), p90, color='#1f77b4')
            ticks, labels = downsample_ticks(x)
            ax.set_xticks(ticks); ax.set_xticklabels(labels, rotation=45, ha='right')
            ax.set_title('90th Percentile Response Time (s) by Transaction')
            ax.set_ylabel('Seconds'); ax.set_ylim(0, 10)
            elements.append(fig_to_image(fig))

        if ps and fl:
            fig, ax = plt.subplots(figsize=(10, 6))
            ax.bar(range(len(x)), ps, label='Pass', color='#4daf4a')
            ax.bar(range(len(x)), fl, bottom=ps, label='Fail', color='#e41a1c')
            ticks, labels = downsample_ticks(x)
            ax.set_xticks(ticks); ax.set_xticklabels(labels, rotation=45, ha='right')
            ax.set_title('Throughput: Pass & Fail counts per Transaction')
            ax.set_ylabel('Count'); ax.legend()
            elements.append(fig_to_image(fig))

        if avg and ps and fl:
            total = (pd.Series(ps) + pd.Series(fl)).tolist()
            fig, ax1 = plt.subplots(figsize=(10, 6))
            ax1.bar(range(len(x)), total, color='#6baed6', alpha=0.6, label='Total Transactions')
            ax2 = ax1.twinx()
            ax2.plot(range(len(x)), avg, color='#ff7f0e', marker='o', label='Average (s)')
            ticks, labels = downsample_ticks(x)
            ax1.set_xticks(ticks); ax1.set_xticklabels(labels, rotation=45, ha='right')
            ax1.set_ylabel('Count'); ax2.set_ylabel('Seconds'); ax2.set_ylim(0, 10)
            ax1.set_title('Transactions vs Count and Average Response Time')
            elements.append(fig_to_image(fig))

        lb = chart_data.get('latency_bands_pie', {})
        if lb.get('labels') and lb.get('values'):
            fig, ax = plt.subplots(figsize=(8, 5))
            ax.pie(lb['values'], labels=lb['labels'], autopct='%1.0f%%')
            ax.set_title('Response Time Distribution (Average)')
            elements.append(fig_to_image(fig, w=500, h=320))

        rc = chart_data.get('resp_code_pie', {})
        if rc.get('labels') and rc.get('values'):
            fig, ax = plt.subplots(figsize=(8, 5))
            ax.pie(rc['values'], labels=rc['labels'], autopct='%1.0f%%')
            ax.set_title('Response Codes')
            elements.append(fig_to_image(fig, w=500, h=320))

    doc.build(elements)
    bio.seek(0)
    return send_file(bio, as_attachment=True,
                     download_name='Performance_Dashboard_Export.pdf',
                     mimetype='application/pdf')


# ----------------------------------------------------------------------------
# Main (for local dev)
# ----------------------------------------------------------------------------
if __name__ == '__main__':
    try:
        for rule in app.url_map.iter_rules():
            methods = ','.join(sorted(rule.methods))
            print(f"Route: {rule.endpoint:25s} Methods: [{methods:20s}] Path: {rule}")
    except Exception:
        pass
    app.run(host='127.0.0.1', port=int(os.environ.get('PORT','5000')), debug=True)
